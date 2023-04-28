from openpyxl import load_workbook
# C##COSMOS.
# All_tables.xlsx

# Replace 'CommonString' with the actual common string you want to remove from sheet names
common_string = '$'

# Replace 'example.xlsx' with the actual name of your Excel file
filename = 'All_tables.xlsx'

# Load the workbook
wb = load_workbook(filename)

# Loop through all sheets in the workbook and rename those that contain the common string
for sheet in wb:
    if common_string in sheet.title:
        new_title = sheet.title.replace(common_string, '')
        if new_title in wb.sheetnames:
            del wb[new_title]
        wb[sheet.title].title = new_title

# Save the modified workbook with a new filename
wb.save('modified_' + filename)


